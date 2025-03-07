import streamlit as st
import openpyxl
import requests
import json
# 设置 DeepSeek API Key 与 URL
API_KEY = "sk-da8664ca636d4d1ba9f1a823f4557a1c"
API_URL = "https://api.deepseek.com/v1/chat/completions"

# 常量定义（请根据实际情况修改）
AUTHOR = "Your Name"
WECHAT_PLATFORM = "Your WeChat Platform"
CURRENT_DATE = "2025-03-07"
# 模型列表（这里使用 deepseek-chat 模型）
model_list = {"deepseek": "deepseek-chat"}


# 处理上传文件函数（这里只处理 Excel 文件）
def process_uploaded_file(uploaded_file):
    try:
        if uploaded_file.type in [
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel"
        ]:
            workbook = openpyxl.load_workbook(uploaded_file, data_only=False)
            return workbook
        else:
            st.error("请上传Excel文件")
            return None
    except Exception as e:
        st.error(f"文件处理失败: {str(e)}")
        return None


# 提取 Excel 工作簿内容（转换为文本格式，默认提取每个 sheet 前50行）
def extract_workbook_content(workbook, max_rows=50):
    content_dict = {}
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        sheet_lines = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            # 转换每个单元格为字符串，以制表符分隔
            row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
            sheet_lines.append(row_text)
        content_dict[sheet_name] = "\n".join(sheet_lines)
    return content_dict


# 侧边栏设置
with st.sidebar:
    st.header("配置参数")
    selected_model = st.selectbox("选择模型", options=list(model_list.keys()), index=0)
    temperature = st.slider("温度参数", 0.0, 1.0, 0.3, 0.1)

    # 新建对话按钮
    if st.button("新建对话"):
        st.session_state.messages = []
        st.session_state.uploaded_content = None
        st.session_state.current_file = None
        st.success("新对话已创建！")

st.title("📑 DeepSeek 智能 Excel 助手 ✨")

st.markdown(
    """<hr style="border:2px solid #FFA07A; border-radius: 5px;">""",
    unsafe_allow_html=True,
)

st.markdown(
    f"""
    <div style='
        text-align: center;
        padding: 15px;
        background: linear-gradient(45deg, #FFD700, #FFA07A);
        border-radius: 15px;
        box-shadow: 0 4px 8px rgba(0,0,0,0.1);
        margin: 20px 0;
    '>
        <h4 style='color: #2F4F4F; margin: 0;'> 作者：{AUTHOR}</h4>
        <p style='color: #800080; margin: 10px 0 0;'>
            微信号：「<strong style='color: #FF4500;'>{WECHAT_PLATFORM}</strong>」
            <br>
            <span style='font-size:14px; color: #4682B4;'>✨ 探索AI的无限可能 ✨</span>
        </p>
    </div>
    """,
    unsafe_allow_html=True,
)

if "first_load" not in st.session_state:
    st.balloons()
    st.session_state.first_load = True

# 文件上传部件（支持 Excel 文件）
uploaded_file = st.file_uploader("上传文档（支持Excel文件）", type=["xlsx", "xls"])
if uploaded_file and uploaded_file != st.session_state.get("current_file"):
    processed_workbook = process_uploaded_file(uploaded_file)
    if processed_workbook:
        st.session_state.uploaded_content = processed_workbook
        st.session_state.current_file = uploaded_file

        # 提取 Excel 内容（各个 sheet 的文本摘要）
        workbook_content = extract_workbook_content(processed_workbook)
        workbook_text = ""
        for sheet, content in workbook_content.items():
            workbook_text += f"Sheet: {sheet}\n{content}\n\n"

        # 构建系统提示，将 Excel sheet 内容直接作为背景输入 DeepSeek API
        system_prompt = f"""
<system>
    [当前日期] {CURRENT_DATE}
    [角色] 您是一名专业的Excel分析助理，擅长找出Excel文件中的人为错误。

    [背景] 
    - 用户上传了文档：{uploaded_file.name}
    - 文档类型：Excel
    - Excel 内容摘要：
{workbook_text}

    [核心任务]
    1. 检查表格中的公式计算是否存在错误或者遗漏数据
    2. 检查表格的求和是否遗漏数据
    3. 不生成Excel中未提供的信息
    4. 保持专业且易懂的语气
    5. 如有疑问，请提示用户提供更多信息

    [交互要求]
    - 保持专业且易懂的语气
    - 关键数据用**加粗**显示
    - 代码块使用```包裹
</system>
        """
        st.session_state.messages = [{"role": "system", "content": system_prompt}]
        st.success(f"Excel文档 {uploaded_file.name} 解析完成！")

# 聊天记录显示（保留系统提示及历史对话）
if "messages" not in st.session_state:
    st.session_state.messages = []

for msg in st.session_state.messages:
    if msg["role"] != "system":
        with st.chat_message(msg["role"]):
            st.markdown(msg["content"])

# 用户输入处理（对话功能，可进一步提问）
if prompt := st.chat_input("请输入问题..."):
    # 添加用户消息
    st.session_state.messages.append({"role": "user", "content": prompt})
    with st.chat_message("user"):
        st.markdown(prompt)

    # 构建API请求的消息上下文（截断历史消息以满足上下文长度要求）
    system_message = st.session_state.messages[0]  # 保留第一条 system 消息
    messages_for_api = [
        {"role": m["role"], "content": m["content"]}
        for m in st.session_state.messages
        if m["role"] != "system"
    ]
    total_length = sum(len(m["content"]) for m in messages_for_api)
    context_length = st.session_state.get("context_length", 16000)
    while total_length > context_length and messages_for_api:
        messages_for_api.pop(0)
        total_length = sum(len(m["content"]) for m in messages_for_api)
    messages_for_api.insert(0, system_message)

    # 构建 DeepSeek API 请求 payload（直接将Excel内容作为背景输入）
    payload = {
        "model": model_list[selected_model],
        "messages": messages_for_api,
        "stream": True,
        "max_tokens": 512,
        "stop": None,
        "temperature": temperature,
        "top_p": 0.7,
        "top_k": 50,
        "frequency_penalty": 0.5,
        "n": 1,
        "response_format": {"type": "text"}
    }

    headers = {
        "Authorization": f"Bearer {API_KEY}",
        "Content-Type": "application/json"
    }

    try:
        # 使用 DeepSeek API 生成流式回复
        with st.chat_message("assistant"):
            response_placeholder = st.empty()
            collected_response = []
            response = requests.post(API_URL, json=payload, headers=headers, stream=True, timeout=30)

            # 新增流式数据处理逻辑
            buffer = ""
            for line in response.iter_lines():
                if line:
                    decoded_line = line.decode("utf-8")

                    # 过滤非数据行
                    if not decoded_line.startswith('data: '):
                        continue

                    # 提取有效JSON数据
                    json_str = decoded_line[6:]  # 去掉"data: "前缀
                    if json_str.strip() == "[DONE]":
                        break

                    try:
                        # 解析JSON并提取关键内容
                        data = json.loads(json_str)
                        if "choices" in data and len(data["choices"]) > 0:
                            delta = data["choices"][0].get("delta", {})
                            content = delta.get("content", "")

                            # 实时更新显示
                            if content:
                                buffer += content
                                response_placeholder.markdown(buffer + "▌")
                    except json.JSONDecodeError:
                        continue

            # 最终显示处理后的内容
            response_placeholder.markdown(buffer)
            st.session_state.messages.append({"role": "assistant", "content": buffer})

    except Exception as e:
        error_msg = f"""
        <error>
            [错误分析]
            API请求失败，可能原因：
            1. 上下文过长（当前：{len(str(messages_for_api))}字符）

            [修正建议]
            请尝试以下操作：
            - 调整上下文长度至16000字符内
            - 重新组织问题表述
            - 新建对话以重试
        </error>
        """
        st.error(error_msg, icon="🚨")
